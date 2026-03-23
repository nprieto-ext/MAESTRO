"""
Script de creation de comptes MyStrow en masse depuis capture/EXCEL.xlsx.
Etape 1 : cree les comptes Firebase Auth + genere les mots de passe
Etape 2 : cree les documents Firestore licenses/{uid} (visibles dans l'admin panel)

Usage : python create_accounts.py
"""
import json
import time
import getpass
import urllib.request
import urllib.error
import openpyxl
from datetime import datetime, timezone
from pathlib import Path

from core import FIREBASE_API_KEY, FIREBASE_PROJECT_ID

EXCEL_PATH       = Path(__file__).parent / "capture" / "EXCEL.xlsx"
EXCEL_OUT_PATH   = Path(__file__).parent / "capture" / "EXCEL_avec_mdp.xlsx"
FIREBASE_CSV     = Path(__file__).parent / "capture" / "firebase_users.csv"
UID_MAP_PATH     = Path(__file__).parent / "capture" / "uid_map.json"

_AUTH_BASE = "https://identitytoolkit.googleapis.com/v1"
_FS_BASE   = (
    f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT_ID}"
    f"/databases/(default)/documents"
)

# ── Mots de passe theme voyage ────────────────────────────────────────────────
TRAVEL_WORDS = [
    "Bali", "Tulum", "Capri", "Lagon", "Oasis", "Safari", "Alize", "Coral",
    "Tahiti", "Cannes", "Maldives", "Tropic", "Breeze", "Sunset", "Voyage",
    "Havana", "Riviera", "Zanzibar", "Mirage", "Horizon", "Lagune", "Pampa",
    "Seychelles", "Samos", "Creta", "Marina", "Palma", "Rhodes", "Marbella",
    "Caribe", "Tropez", "Papaya", "Ibiza", "Malaga", "Mykonos", "Santorini",
    "Bodrum", "Cancun", "Nassau", "Reunion", "Borneo", "Phuket", "Lombok",
    "Aruba", "Madeira", "Azores", "Corsica", "Sardinia", "Sicilia", "Antalya",
    "Aqaba", "Hurghada", "Marsa", "Djerba", "Agadir", "Maputo", "Noumea",
    "Moorea", "Rarotonga", "Vanuatu", "Curaçao", "Essaouira",
]

def _make_password(club_name: str, index: int) -> str:
    word   = TRAVEL_WORDS[index % len(TRAVEL_WORDS)]
    number = 100 + (index * 37 + len(club_name)) % 900
    return f"{word}{number}"


# ── Firebase helpers ──────────────────────────────────────────────────────────

def _post(url, payload):
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json"}, method="POST"
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read().decode())

def _firebase_sign_up(email, password):
    url  = f"{_AUTH_BASE}/accounts:signUp?key={FIREBASE_API_KEY}"
    resp = _post(url, {"email": email, "password": password, "returnSecureToken": True})
    return resp["localId"], resp["idToken"]

def _firebase_sign_in(email, password):
    url  = f"{_AUTH_BASE}/accounts:signInWithPassword?key={FIREBASE_API_KEY}"
    resp = _post(url, {"email": email, "password": password, "returnSecureToken": True})
    return resp["localId"], resp["idToken"]

def _load_firebase_csv():
    """Charge le CSV exporte depuis Firebase Console (Authentication > Export users).
    Retourne un dict email -> uid."""
    if not FIREBASE_CSV.exists():
        return {}
    import csv
    mapping = {}
    with open(FIREBASE_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = (row.get("Email Address", "") or row.get("Identifier", "")).strip().lower()
            uid   = (row.get("User UID", "") or row.get("uid", "")).strip()
            if email and uid:
                mapping[email] = uid
    return mapping

def _firebase_lookup_uid(email, admin_token):
    """Recupere l'uid d'un compte existant via le token admin (sans connaitre le mot de passe)."""
    url     = f"{_AUTH_BASE}/accounts:lookup?key={FIREBASE_API_KEY}"
    payload = json.dumps({"email": [email]}).encode()
    req     = urllib.request.Request(
        url, data=payload,
        headers={
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {admin_token}",
        },
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        resp = json.loads(r.read().decode())
    users = resp.get("users", [])
    if not users:
        raise ValueError(f"Aucun utilisateur trouve pour {email}")
    return users[0]["localId"]

def _to_fs(value):
    """Convertit une valeur Python en champ Firestore."""
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        return {"doubleValue": value}
    if isinstance(value, str):
        return {"stringValue": value}
    if isinstance(value, list):
        return {"arrayValue": {"values": [_to_fs(v) for v in value]}}
    return {"stringValue": str(value)}

def _create_license_doc(uid, email, admin_token):
    """Cree le document licenses/{uid} dans Firestore."""
    fields = {
        "email":       _to_fs(email),
        "plan":        _to_fs("license"),
        "expiry_utc":  _to_fs(datetime(2026, 12, 31, tzinfo=timezone.utc).timestamp()),
        "created_utc": _to_fs(datetime.now(timezone.utc).timestamp()),
        "machines":    _to_fs([]),
    }
    url     = f"{_FS_BASE}/licenses/{uid}"
    payload = json.dumps({"fields": fields}).encode()
    req     = urllib.request.Request(
        url, data=payload,
        headers={
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {admin_token}",
        },
        method="PATCH",
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        r.read()

def _find_uid_in_firestore(email, admin_token):
    """Cherche l'UID dans la collection licenses en filtrant par email.
    Utile quand le compte Firebase Auth existe mais que l'UID est inconnu."""
    url   = (
        f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT_ID}"
        f"/databases/(default)/documents:runQuery"
    )
    query = {
        "structuredQuery": {
            "from":  [{"collectionId": "licenses"}],
            "where": {
                "fieldFilter": {
                    "field": {"fieldPath": "email"},
                    "op":    "EQUAL",
                    "value": {"stringValue": email},
                }
            },
            "limit": 1,
        }
    }
    payload = json.dumps(query).encode()
    req     = urllib.request.Request(
        url, data=payload,
        headers={
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {admin_token}",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        results = json.loads(r.read().decode())
    for result in results:
        doc = result.get("document")
        if doc:
            # "name" = projects/.../databases/(default)/documents/licenses/{uid}
            return doc["name"].split("/")[-1]
    return None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Connexion admin (necessaire pour ecrire dans Firestore)
    print("Connexion admin requise pour creer les documents Firestore.")
    admin_email    = input("Email admin : ").strip()
    admin_password = getpass.getpass("Mot de passe admin : ")
    try:
        _, admin_token = _firebase_sign_in(admin_email, admin_password)
        print(f"Connecte en tant que {admin_email}\n")
    except Exception as e:
        print(f"Echec connexion admin : {e}")
        return

    uid_csv_map = _load_firebase_csv()
    if uid_csv_map:
        print(f"CSV Firebase charge : {len(uid_csv_map)} utilisateurs\n")

    uid_manual_map = {}
    if UID_MAP_PATH.exists():
        with open(UID_MAP_PATH, encoding="utf-8") as f:
            raw = json.load(f)
        uid_manual_map = {k.lower(): v for k, v in raw.items() if v != "COLLE_UID_ICI"}
        if uid_manual_map:
            print(f"uid_map.json charge : {len(uid_manual_map)} UIDs manuels\n")

    read_path = EXCEL_OUT_PATH if EXCEL_OUT_PATH.exists() else EXCEL_PATH
    wb   = openpyxl.load_workbook(read_path)
    ws   = wb.active
    rows = list(ws.iter_rows(min_row=2, min_col=1, max_col=4, values_only=False))

    if ws.cell(1, 4).value is None:
        ws.cell(1, 4).value = "Mot de passe"

    total = len(rows)
    print(f"{total} clubs a traiter\n")

    ok_count  = 0
    err_count = 0

    for i, row in enumerate(rows):
        club  = row[1].value or ""
        email = (row[2].value or "").strip()
        if not email:
            print(f"[{i+1:02d}] SKIP  (pas d'email)")
            continue

        # Reutiliser le mot de passe deja genere ou en creer un nouveau
        existing_pwd = row[3].value
        password = existing_pwd if existing_pwd else _make_password(club, i)

        # Etape 1 : creer ou recuperer le compte Auth
        uid = None
        # UID deja connu ? On saute completement les appels Firebase Auth
        known_uid = uid_manual_map.get(email.lower()) or uid_csv_map.get(email.lower())
        if known_uid:
            uid = known_uid
            auth_status = "Auth manuel"
        else:
            try:
                uid, _ = _firebase_sign_up(email, password)
                auth_status = "Auth cree"
            except urllib.error.HTTPError as e:
                body = e.read().decode(errors="replace")
                if "EMAIL_EXISTS" in body:
                    # 1) Essai connexion avec le mot de passe genere
                    try:
                        uid, _ = _firebase_sign_in(email, password)
                        auth_status = "Auth existant"
                    except Exception:
                        pass
                    # 2) Fallback : chercher l'UID dans Firestore (doc licenses existant)
                    if not uid:
                        try:
                            uid = _find_uid_in_firestore(email, admin_token)
                            if uid:
                                auth_status = "UID via Firestore"
                                # Persist dans uid_map.json pour les prochains runs
                                uid_manual_map[email.lower()] = uid
                                raw_map = {}
                                if UID_MAP_PATH.exists():
                                    with open(UID_MAP_PATH, encoding="utf-8") as f:
                                        raw_map = json.load(f)
                                raw_map[email] = uid
                                with open(UID_MAP_PATH, "w", encoding="utf-8") as f:
                                    json.dump(raw_map, f, indent=2, ensure_ascii=False)
                        except Exception:
                            pass
                    if not uid:
                        print(f"[{i+1:02d}] SKIP {email} : mot de passe inconnu, UID introuvable")
                        err_count += 1
                        continue
                else:
                    print(f"[{i+1:02d}] ERREUR Auth {email} : {body[:80]}")
                    err_count += 1
                    continue
            except Exception as e:
                print(f"[{i+1:02d}] ERREUR {email} : {e}")
                err_count += 1
                continue

        # Etape 2 : creer le document Firestore licenses/{uid}
        try:
            _create_license_doc(uid, email, admin_token)
            fs_status = "Firestore OK"
        except Exception as e:
            fs_status = f"Firestore ERREUR : {e}"
            err_count += 1

        # Sauvegarder le mot de passe dans le fichier
        row[3].value = password
        wb.save(EXCEL_OUT_PATH)

        print(f"[{i+1:02d}] {auth_status:<16} | {fs_status:<14} | {email:<45}  {password}")
        ok_count += 1
        time.sleep(0.15)

    wb.save(EXCEL_OUT_PATH)
    print(f"\nTermine : {ok_count} traites - {err_count} erreurs")
    print(f"Fichier : {EXCEL_OUT_PATH}")


if __name__ == "__main__":
    main()
